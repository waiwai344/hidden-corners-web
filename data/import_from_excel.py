from __future__ import annotations

import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = ROOT.parent
DB_PATH = ROOT / "data" / "hidden-corners.sqlite"
SCHEMA_PATH = ROOT / "data" / "schema.sql"
EXCEL_PATH = SOURCE_ROOT / "数据库Excel版.xlsx"


def normalize_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return (date(1899, 12, 30) + timedelta(days=int(value))).isoformat()
    return str(value or "")[:10]


def rows(workbook, sheet_name):
    sheet = workbook[sheet_name]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    for record in sheet.iter_rows(min_row=2, values_only=True):
        yield dict(zip(header, record))


def insert_many(conn, table, columns, source_rows, mapper):
    placeholders = ", ".join("?" for _ in columns)
    column_sql = ", ".join(columns)
    values = [[mapper(column, row) for column in columns] for row in source_rows]
    conn.executemany(f"INSERT INTO {table} ({column_sql}) VALUES ({placeholders})", values)


def text(value):
    return str(value or "").strip()


def number(value):
    return int(value)


def main():
    workbook = load_workbook(EXCEL_PATH, data_only=True)
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))

    insert_many(
        conn,
        "Spaces",
        ["SpaceID", "SpaceName", "City", "Description", "Address", "Longitude", "Latitude"],
        rows(workbook, "Spaces"),
        lambda column, row: float(row[column]) if column in ["Longitude", "Latitude"] else number(row[column]) if column == "SpaceID" else text(row[column]),
    )
    insert_many(
        conn,
        "Categories",
        ["CategoryID", "CategoryName", "CategoryDesc"],
        rows(workbook, "Categories"),
        lambda column, row: number(row[column]) if column == "CategoryID" else text(row[column]),
    )
    insert_many(
        conn,
        "SpaceCategories",
        ["SCID", "SpaceID", "CategoryID"],
        rows(workbook, "SpaceCategories"),
        lambda column, row: number(row[column]),
    )
    insert_many(
        conn,
        "NomadCommunities",
        ["CommunityID", "CommunityName", "Province", "City", "Description", "Capacity", "MonthlyPrice"],
        rows(workbook, "NomadCommunities"),
        lambda column, row: number(row[column]) if column in ["CommunityID", "Capacity", "MonthlyPrice"] else text(row[column]),
    )
    insert_many(
        conn,
        "Users",
        ["UserID", "Username", "PasswordHash", "Gender", "BirthDate", "HomeCity", "UserType", "RegisterDate"],
        rows(workbook, "Users"),
        lambda column, row: number(row[column]) if column == "UserID" else normalize_date(row[column]) if column in ["BirthDate", "RegisterDate"] else text(row[column]),
    )
    insert_many(
        conn,
        "Reviews",
        ["ReviewID", "UserID", "SpaceID", "Rating", "Content", "VisitDate"],
        rows(workbook, "Reviews"),
        lambda column, row: number(row[column]) if column in ["ReviewID", "UserID", "SpaceID", "Rating"] else normalize_date(row[column]) if column == "VisitDate" else text(row[column]),
    )
    insert_many(
        conn,
        "Favorites",
        ["FavoriteID", "UserID", "SpaceID", "ActionType", "ActionDate"],
        rows(workbook, "Favorites"),
        lambda column, row: number(row[column]) if column in ["FavoriteID", "UserID", "SpaceID"] else normalize_date(row[column]) if column == "ActionDate" else text(row[column]),
    )
    insert_many(
        conn,
        "Activities",
        ["ActivityID", "SpaceID", "ActivityName", "ActivityDate", "PushLink"],
        rows(workbook, "Activities"),
        lambda column, row: number(row[column]) if column in ["ActivityID", "SpaceID"] else normalize_date(row[column]) if column == "ActivityDate" else text(row[column]),
    )

    conn.commit()
    for table in ["Spaces", "Categories", "SpaceCategories", "NomadCommunities", "Users", "Reviews", "Favorites", "Activities"]:
        count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        print(f"{table}: {count}")
    conn.close()


if __name__ == "__main__":
    main()
